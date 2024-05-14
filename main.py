import imaplib
import os
import re
import pandas as pd
import openpyxl

# Substitua pelos seus dados
username = "email@gmail.com" #email aqui
password = "senha" #senha aqui

# Substitua pelos termos desejados
assunto = "Comunicado de Pagamento"
termo_data = "data de"
termo_nota = "Num.Nfe"
termo_valor = "Valor" 

def extrair_informacao(corpo_mensagem, termo):
    resultado = re.search(rf"{termo}: (.+)", corpo_mensagem)
    if resultado:
        return resultado.group(1).strip()
    else:
        return None

imap_server = "imap.gmail.com"
imap_port = 993

connection = imaplib.IMAP4_SSL(imap_server, imap_port)
connection.login(username, password)

# Selecionar Caixa de Entrada e Pesquisar Emails
caixa_de_entrada = "INBOX"
consulta = f"(FROM 'FINANCEIRO.AVB@ferroeste.com.br' OR FROM 'FINANCEIRO.AVB@ferroeste.com.br') SUBJECT '{assunto}'"

connection.select(caixa_de_entrada)
connection.search(None, consulta)

dados = []

status, dados_resposta = connection.search(None, consulta)
id_mensagens = dados_resposta[0].split(b' ')

for id_mensagem in id_mensagens:
    id_mensagem = int(id_mensagem.decode())

    status, dados_mensagem = connection.fetch(id_mensagem, '(RFC822)')
    corpo_mensagem = dados_mensagem[0][1].decode('utf-8')

    data = extrair_informacao(corpo_mensagem, termo_data)
    numero_nota = extrair_informacao(corpo_mensagem, termo_nota)
    valor_nota = extrair_informacao(corpo_mensagem, termo_valor)

    dados.append([data, numero_nota, valor_nota])

# Criar Planilha do Excel e Salvar Dados
arquivo_excel = "Dados_Notas_Fiscais.xlsx"
planilha = openpyxl.load_workbook(arquivo_excel, data_only=True) if os.path.exists(arquivo_excel) else openpyxl.Workbook()

nome_planilha = "Notas Fiscais"
planilha_nova = planilha.create_sheet(nome_planilha)

planilha_nova.append(['Data Pagamento', 'Número da Nota', 'Valor da Nota'])
for i in range(len(dados)):
    planilha_nova.append(dados[i].tolist())

planilha.save(arquivo_excel)